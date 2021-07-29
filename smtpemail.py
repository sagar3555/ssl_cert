import subprocess
import openpyxl
import os
import datetime


from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from DBConnection_Sagar import Base,cert_data_sagar
engine = create_engine(r'sqlite:///E:\Scripts\Project_Cert\certdetail_sagar.db')
Base.metadata.bind = engine
DBSession = sessionmaker(bind=engine)





def Get_email_addr(extid) :
    fh = open("get_email_address.ps1", "w+")
    fh.write("import-module activedirectory \n")
    fh.write("get-aduser -filter{samaccountname -eq "+ "'"+extid+"'"+"} -properties mail | Select -expandproperty mail")

    fh.close()
    ch = subprocess.Popen([r'C:\WINDOWS\system32\WindowsPowerShell\v1.0\powershell.exe', r'E:\4-FlaskForms\get_email_address.ps1'], shell=False,
                          stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = ch.communicate()
    print((stdout.decode()),len(stdout.decode()))
    if len(stdout.decode()) > 0:
        return str(stdout.decode())
    else:
        return "cominfo@ica.se"


#Get_email_addr("extsdd")

def Get_Technical_Owner (ci_name) :
    print (ci_name)
    if ci_name.upper()!= "PKI"  or  ci_name.upper == "COMINFO":
        input_file_path = os.path.join("E:\\" ,"4-FlaskForms","Technical Owner Details.xlsx")
        wb = openpyxl.load_workbook(filename=input_file_path)
        ws = wb.get_sheet_by_name('Sheet1')
        rowmax = ws.max_row
        flag = False
        for row in range  (2, rowmax+1) :
            if  ws.cell(row=row, column = 2).value != None  :
                if  ws.cell(row=row, column = 1).value.strip().upper()  == ci_name.upper() :
                    technical_owner = ws.cell(row=row, column = 2).value.strip().upper()
                    flag = True
                    break

        if flag == True :
            technical_owner_email  = Get_email_addr(technical_owner)
        else  :
            technical_owner_email = "cominfo@ica.se"
    else :
        technical_owner_email = "cominfo@ica.se"

    return technical_owner_email



def First_mail (cert_name , expiration_date,technical_owner) :

    SMTPserver = 'smtp-relay.ica.ia-hc.net'
    sender =     'cominfo@ica.se'
    destination = [technical_owner]

    # typical values for text_subtype are plain, html, xml
    text_subtype = 'plain'


    content= "PFB details for Certificate expiry \n EXpiration Date :"+str(expiration_date)+"\n Link for Renewal web Form :  https://10.107.198.101:80/ (best Viewed In chrome) "

    subject="Certificate Expiry Notification 1: "+cert_name

    import sys
    import os
    import re

    #from smtplib import SMTP_SSL as SMTP       # this invokes the secure SMTP protocol (port 465, uses SSL)
    from smtplib import SMTP                  # use this for standard SMTP protocol   (port 25, no encryption)

    # old version
    # from email.MIMEText import MIMEText
    from email.mime.text import MIMEText


    msg = MIMEText(content, text_subtype)
    msg['Subject']=       subject
    msg['From']   = sender # some SMTP servers will do this automatically, not all

    conn = SMTP(SMTPserver)
    conn.set_debuglevel(False)
    #conn.login(USERNAME, PASSWORD)

    conn.sendmail(sender, destination, msg.as_string())
    conn.quit()

#First_mail("a","2018-02-03","sagar.debadwar@ica.se")



def Second_Notification () :
    session = DBSession()
    cert_d = session.query(cert_data_sagar).all()
    for cert in cert_d:
        if  25 < cert.expiration_date - datetime.datetime.now() <=30 and  cert.renewal_status.lower()== "initial" :
            technical_owner  = cert. technical_owner
            SMTPserver = 'smtp-relay.ica.ia-hc.net'
            sender = 'cominfo@ica.se'
            destination = [technical_owner]
            text_subtype = 'plain'

            content = "PFB details for Certificate expiry \n EXpiration Date :" + str(
                cert.expiration_date) + "\n Link for Renewal web Form :  https://10.107.198.101:80/ (best Viewed In chrome) "

            subject = "Certificate Expiry Notification 2: " + cert.cert_name

            import sys
            import os
            import re

            # from smtplib import SMTP_SSL as SMTP       # this invokes the secure SMTP protocol (port 465, uses SSL)
            from smtplib import SMTP  # use this for standard SMTP protocol   (port 25, no encryption)

            # old version
            # from email.MIMEText import MIMEText
            from email.mime.text import MIMEText

            msg = MIMEText(content, text_subtype)
            msg['Subject'] = subject
            msg['From'] = sender  # some SMTP servers will do this automatically, not all

            conn = SMTP(SMTPserver)
            conn.set_debuglevel(False)
            # conn.login(USERNAME, PASSWORD)

            conn.sendmail(sender, destination, msg.as_string())
            conn.quit()
            cert.renewal_status= "second notification sent"
    session.close()



def Third_Notification ():
    session = DBSession()
    cert_d = session.query(cert_data_sagar).all()
    for cert in cert_d:
        if  cert.expiration_date - datetime.datetime.now() <= 10 and cert.renewal_status.lower() != "details received for renewal":
            technical_owner = cert.technical_owner
            SMTPserver = 'smtp-relay.ica.ia-hc.net'
            sender = 'cominfo@ica.se'
            destination = [technical_owner]
            text_subtype = 'plain'

            content = "PFB details for Certificate expiry \n EXpiration Date :" + str(
                cert.expiration_date) + "\n Link for Renewal web Form :  https://10.107.198.101:80/ (best Viewed In chrome) "

            subject = "Certificate Expiry Notification 3: " + cert.cert_name

            import sys
            import os
            import re

            # from smtplib import SMTP_SSL as SMTP       # this invokes the secure SMTP protocol (port 465, uses SSL)
            from smtplib import SMTP  # use this for standard SMTP protocol   (port 25, no encryption)

            # old version
            # from email.MIMEText import MIMEText
            from email.mime.text import MIMEText

            msg = MIMEText(content, text_subtype)
            msg['Subject'] = subject
            msg['From'] = sender  # some SMTP servers will do this automatically, not all

            conn = SMTP(SMTPserver)
            conn.set_debuglevel(False)
            # conn.login(USERNAME, PASSWORD)

            conn.sendmail(sender, destination, msg.as_string())
            conn.quit()
            cert.renewal_status = "final notification sent"
            session.commit()

    session.close()


